import os, re, html, string, datetime
import boto3
import pandas as pd
from io import BytesIO
from random import choice
from openpyxl import load_workbook
from django.http import HttpResponse
from collections import defaultdict


from urllib.parse import quote,unquote
from django.conf import settings
from django.utils.html import escape

## (메모) 
## is_valid_info : 입력된 가입 정보가 유효한지 확인
## @param user_info (List 타입) : 입력된 가입 정보 묶음
## @return 
## - Set 타입 : 상태 메시지 ('status' : 유효 여부 - (True=유효함, False=유효하지 않음), 'msg' : 상태 메시지)

def is_valid_info(user_info):
    user_id, user_pw, user_pwCheck, user_name, user_email, user_hp, user_type = user_info
    ## 
    if is_null(user_info):
        return {'status' : False, 'msg' : '모든 값을 입력해주세요.'}
    if is_validID(user_id):
        return {'status' : False, 'msg' : '아이디는 반드시 4~12자 영문/숫자로 구성되어야 합니다.'}
    if is_validPW(user_pw):
        return {'status' : False, 'msg' : '비밀번호는 반드시 10자 이상의 최소하나씩의 영문/숫자/특수문자로 구성되어야 합니다.'}    
    if is_validHP(user_hp):     
        return {'status':False ,'msg': '유효하지 않은 핸드폰 번호 형식입니다.'}
    if user_pw != user_pwCheck :
        return { 'status': False, 'msg': '비밀번호가 일치하지 않습니다.'}
    if not is_validType(user_type) :
        return { 'status': False, 'msg': '회원 유형은 평가자 또는 OEM 만 존재합니다.'}    
    return {'status': True, 'msg' : '성공'}


## (메모)
## is_null : 입력된 정보가 없는지 확인
## @param 
## - user_info : 가입 정보
## @return
## - Bool 타입 : 입력된 가입정보 없다면 True, 있다면 False 
def is_null(user_info):
    return None in user_info


## (메모)
## is_validID : 입력된 ID가 유효한지 확인
## @param 
## - user_id : 가입 ID
## @return
## - Bool 타입 : ID가 유효하다면 True, 아니라면 False 
def is_validID(user_id):
    patt = re.compile('^[0-9|a-z|A-Z]{4,12}$')
    m = patt.match(user_id)
    return m is None


## (메모)
## is_validPW : 입력된 PW가 유효한지 확인
## @param 
## - user_pw : 가입 pw
## @return
## - Bool 타입 : pw가 유효하다면 True, 아니라면 False 
def is_validPW(user_pw):
    patt = re.compile('^(?=.*[A-Za-z])(?=.*\d)(?=.*[@$!%*#?&])[A-Za-z\d@$!%*#?&]{10,}$')
    m = patt.match(user_pw)
    return m is None

## (메모)
## is_validHP : 입력된 전화번호가 유효한지 확인
## @param 
## - user_hp : 가입 전화번호
## @return
## - Bool 타입 : 전화번호가 유효하다면 True, 아니라면 False 
def is_validHP(user_hp):
    patt = re.compile('^\s*(010|011)([0-9]{4}){2}\s*$')
    m = patt.match(user_hp)
    return m is None

## (메모)
## is_validType : 입력된 회원 유형이 유효한지 확인
## @param 
## - user_type : 회원 유형
## @return
## - Bool 타입 : 회원 유형이 유효(OEM-OEM 또는 EVL-평가자)하다면 True, 아니라면 False
def is_validType(user_type):
    return user_type== "OEM" or user_type== "EVL"  

## (메모)
## is_email_vrfied : 이메일 인증을 마친 회원인지 확인
## @param 
## - request : 가입 요청 정보
## @return
## - Set 타입 : 상태 메시지 ('status' : 유효 여부 - (True=유효함, False=유효하지 않음), 'msg' : 상태 메시지)
def is_email_vrfied(request):
    print(request.session['email_verify'].keys())
    print(request.session['email_verify']['is_vrfd'])
    if 'email_verify' not in list(request.session.keys()):
        return {'status': False,'msg': '먼저 이메일 인증을 하세요.'}
        ##render(request, 'signup_backend.html', {'error': '먼저 문자인증을 하세요.'})
    if 'is_vrfd' not in list(request.session['email_verify'].keys()) or not request.session['email_verify']['is_vrfd'] :
        return {'status' : False, 'msg': '인증번호가 검증되지 않았습니다.'}
    return {'status' : True, 'msg' : '성공'}

## (메모)
## is_expired : 이메일 인증 번호의 유효기간이 지났는지 확인
## @param 
## - str_time : 인증번호가 유효기한 
## @return
## - Bool 타입 : 인증번호의 유효기한이 지났다면 True, 아니라면 False 
def is_expired(str_time):
    due = datetime.datetime.strptime(str_time,"%Y-%m-%d %H:%M:%S")
    now = datetime.datetime.now().replace(microsecond=0)
    return due < now

## (메모)
## set_expire : 이메일 인증 번호의 유효기간 설정
## @param 
## @return
## - 시간 정보 : 인증번호의 만료 시간 (3분) 
def set_expire():
    due = datetime.datetime.now() + datetime.timedelta(seconds=60*3) ## 인증번호는 최대 3분 유효함. 
    return due.strftime("%Y-%m-%d %H:%M:%S")

## (메모)
## random_code_generator : 이메일 인증 번호 생성
## @param 
## @return
## - 시간 정보 : 6자리 무작위 인증번호 생성
def random_code_generator():
    randoms =  ''.join(choice(string.digits) for _ in range(6))
    return randoms


## (메모)
## read_excel_with_merge : 엑셀 파일에서 병합된 셀 정보를 추출
## @param
## - excel_file : 읽어들인 엑셀 파일 
## @return
## - Dataframe 타입 : 엑셀에서 읽어들인 Dataframe
## - List 타입 : 병합된 셀 정보 (시작행, 시작열, 병합된 행 길이, 병합된 열 길이)
def read_excel_with_merge(excel_file):  
    workbook = load_workbook(excel_file)
    sheet = workbook.active

    df = pd.DataFrame(sheet.values)

    # (메모) 엑셀 데이터를 읽어오기
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))
    
    # (메모) 병합된 셀 정보 가져오기
    merge_info = []
    for merge_range in sheet.merged_cells.ranges:
        start_cell = merge_range.min_row - 1, merge_range.min_col - 1
        merge_info.append({
            "start_row": start_cell[0],
            "start_col": start_cell[1],
            "rowspan": merge_range.max_row - merge_range.min_row + 1,
            "colspan": merge_range.max_col - merge_range.min_col + 1,
        })
    
    return df, merge_info


## (메모)
## read_excel_without_merge : 엑셀 파일에서  정보를 추출
## @param
## - excel_file : 읽어들인 엑셀 파일 
## @return
## - Dataframe 타입 : 엑셀에서 읽어들인 Dataframe
## - List 타입 : 병합된 셀 정보 (시작행, 시작열, 병합된 행 길이, 병합된 열 길이)
def read_excel_without_merge(excel_file):  
    workbook = load_workbook(excel_file)
    sheet = workbook.active

    # 병합된 셀 값 채우기
    for merged_range in sheet.merged_cells.ranges:
        merged_cells = list(merged_range.cells)
        value = sheet.cell(*merged_cells[0]).value
        for cell in merged_cells:
            sheet.cell(*cell).value = value

    # 엑셀 데이터를 DataFrame으로 변환
    data = sheet.values
    columns = next(data)  # 첫 번째 줄을 컬럼으로 사용
    df = pd.DataFrame(data, columns=columns)
    print(df)
    return df.to_dict(orient='records')


## (메모)
## alert_and_redirect : 알림(alert)을 띄우고 다른 url로 리다이렉팅
## @param
## - message : 알림을 통해 띄우고자 하는 메시지
## - redirect_url : 리다이렉팅 될 url
## @return
## - Http 응답 : 알림 & 리다이렉팅을 수행하는 Http 응답
def alert_and_redirect(message, redirect_url):
    safe_message = escape(message)  # 메시지를 HTML 이스케이프 처리
    safe_redirect_url = escape(redirect_url)  # URL도 안전하게 이스케이프 처리

    html = f"""
        <script>
            alert("{safe_message}");
            window.location.href = "{safe_redirect_url}";
        </script>
    """
    return HttpResponse(html)


def search_all_files(s3_client):
    files_info = []
    paginator = s3_client.get_paginator('list_objects_v2')
    
    for page in paginator.paginate(Bucket=settings.AWS_STORAGE_BUCKET_NAME, Prefix='eval_state'):
        if 'Contents' in page:
                for obj in page['Contents']:
                    if obj['Size'] != 0:
                        file_key = obj['Key']                # 증거물 파일 이름
                                                        
                        # 사용자 정의 메타데이터 조회
                        try:
                            download_url = get_download_link(s3_client,file_key)
                            head_response = s3_client.head_object(Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=file_key)
                            custom_metadata = head_response.get('Metadata', {})  # 사용자 정의 메타데이터
                            files_info.append({
                                'download_url' : download_url,
                                'created_time' : custom_metadata['created_time'],
                                'filename' : os.path.basename(file_key),
                                'original_name' : unquote(custom_metadata['original_name']),
                                'username' : custom_metadata['username'],
                                'oem' :custom_metadata['oem'],
                                'key' : file_key,
                                'state' : custom_metadata['state']
                                })

                        except Exception as e:
                            raise

    return files_info



def get_paginater(page_number,page_data_list):
    page_count=len(page_data_list)//15  ## 전체 페이지 갯수
    
    if len(page_data_list)%15:
        page_count = page_count + 1
    
    current_page_number = page_number   # URL에서 'page' 파라미터 가져오기
    
    if not(current_page_number is None):
        current_page_number = int(page_number)
    else:
        current_page_number = 1
    
    page_data = page_data_list[0+(current_page_number-1)*15:0+current_page_number*15]

    start_page = page_number-1
    if start_page == 0:
        start_page = 1
    final_page = page_number+1
    if final_page == page_count+1 :
        final_page = page_count

    page_range = list(range(start_page,final_page+1))

    paination_obj = {
        "page_number": page_number,            ## 현재 조회하고자 하는 페이지
        "page_count": page_count,              ## 페이지 전체 갯수
        "one_page_count" : 15,                  ## 한 페이지내 전체 갯수
        "page_data" : page_data,
        "prev_page" : page_number-1,
        "next_page" : page_number+1,
        "page_range" : page_range,
    }

    return paination_obj


def search_user_files(s3_client, username):
    files_info_list = []
    ##folder_name = username
    ##escaped_folder_name = quote(folder_name.strip('/'))
    ##pending_path= 'eval_pending/' + escaped_folder_name
    ##complete_path= 'eval_complete/' + escaped_folder_name

    try:
        ## 평가 대기 중인 파일의 목록
        files_info_list = search_all_files(s3_client)
        filtered_files_info = [item for item in files_info_list if item.get("username") == username]
        return filtered_files_info
            
    except Exception as e:
        raise
    

    
def get_download_link(s3_client,filename):
    download_link = ""
    try:
        download_link = s3_client.generate_presigned_url(
            'get_object',
            Params={'Bucket': settings.AWS_STORAGE_BUCKET_NAME, 'Key': filename},
            ExpiresIn=3600
        )
        return html.unescape(download_link)
    except Exception as e:
        raise

def get_current_checklist(s3_client):
     ## S3 버킷 접근 코드

    try:
        response = s3_client.get_object(Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key='current_eval_list/checklist.xlsx')
        ## (메모) S3에서 엑셀 파일 읽기
        file_content = response['Body'].read()
        excel_file = BytesIO(file_content)

        ## 병합셀 있는 버전 (추후 시간 남으면 구현)
        ##df, merge_data = read_excel_with_merge(excel_file)

        ## 병합셀 없는 버전
        df = pd.read_excel(excel_file, sheet_name=0)
        df.fillna(method='ffill', inplace=True)
        table_data_temp = df.to_dict(orient='records')  # DataFrame을 리스트로 변환 (템플릿에서 사용할 수 있도록) [{column1: value1, column2: value2}, ...]
        columns = df.columns.tolist() # 컬럼 이름도 전달
        columns = columns[:len(columns)-1]

        table_data = [
            {key.replace(" ", "_"): value for key, value in item.items()}
            for item in table_data_temp
        ]

        return table_data, columns
    except Exception as e:
        raise

def get_Etag(s3_client,file_key):
    etag= ""
    try:
        response = s3_client.head_object(Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=file_key)
        etag = response.get('ETag', None)
        return etag  # ETag is returned in the response metadata
    except Exception as e:
        raise

def identify_submit_checklist(filename):
    s3_client = boto3.client('s3',
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                region_name=settings.AWS_REGION,
            )

    try:
        head_response = s3_client.head_object(Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=filename)
        custom_metadata = head_response.get('Metadata', {})  # 사용자 정의 메타데이터
        return custom_metadata['checklist_hash']
        
    except Exception as e:
        raise


def get_file_metadata(s3_client,file_key,metadata):
    try:
        head_response = s3_client.head_object(Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=file_key)
        custom_metadata = head_response.get('Metadata', {})  # 사용자 정의 메타데이터
        custom_metadata['original_name'] = unquote(custom_metadata['original_name'])
        return custom_metadata[metadata]
    except Exception as e:
        raise

def has_such_file_with_key(s3_client,file_key):
    try:
        s3_client.head_object(Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=file_key)  # Check file existence
        return True  # File exists
    except s3_client.exceptions.ClientError as e:
        # If the error is 404 (Not Found), the file does not exist
        if e.response['Error']['Code'] == "404":
            return False
        else:
            raise  # Re-raise any other exceptions
    except (NoCredentialsError, PartialCredentialsError):
        return False


def upload_submission_file(s3_client,username,filedata,current_checklist_hash) : 
    try:
        # S3에 업로드
        original_name = filedata.name
        nowtime = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        filedata.name = username+'_'+nowtime+ '.zip'
        s3_file_path = f"eval_state/{filedata.name}"  # S3 경로 설정
            ##s3_client.put_object(Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=s3_file_path)

        
        s3_client.upload_fileobj(filedata, settings.AWS_STORAGE_BUCKET_NAME, s3_file_path, ExtraArgs={
                    'Metadata': {
                        'checklist_hash': current_checklist_hash,
                        'created_time' : nowtime, 
                        'original_name' : quote(original_name),
                        'username' : username,
                        ##'oem' : (request.user.email)
                        'oem' : 'Hyundai',
                        'state' : 'pend'
                    },
                    'ContentType': 'application/zip',
                    })
            # 성공적으로 업로드된 경우 응답
    except Exception as e:
        raise


def modify_submission_metadata(s3_client,file_key,metadata,value):
    try:
        head_response = s3_client.head_object(Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=file_key)
        current_metadata = head_response.get('Metadata', {})

        # 새로운 메타데이터 설정 (기존 메타데이터에 덮어쓰기)
        # 복사 후 'key2' 값만 변경
        updated_metadata = {**current_metadata, metadata: value}
        # 객체 복사로 메타데이터 업데이트
        s3_client.copy_object(
            Bucket=settings.AWS_STORAGE_BUCKET_NAME,
            CopySource={'Bucket': settings.AWS_STORAGE_BUCKET_NAME, 'Key': file_key},
            Key=file_key,
            Metadata=updated_metadata,
            MetadataDirective='REPLACE'
        )
        return True
    except Exception as e:
        raise



def get_completion_ratio(result_list):
    # 그룹화 및 비율 계산
    grouped_data = defaultdict(list)
    for item in result_list:
        grouped_data[item["category"]].append(item["passfail"])

    # 각 그룹의 P/F 비율 계산
    category_ratios = {}
    for category, passfail_list in grouped_data.items():
        total = len(passfail_list)
        p_count = passfail_list.count("P")
        f_count = passfail_list.count("F")
        category_ratios[category] = p_count / total if total > 0 else 0
        

    return category_ratios

